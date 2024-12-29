from flask import Flask, request, redirect, send_file
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# ตรวจสอบว่าไฟล์ Excel มีอยู่แล้วหรือไม่ ถ้าไม่มี ให้สร้างใหม่
EXCEL_FILE = 'data.xlsx'
if not os.path.exists(EXCEL_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Messages"
    sheet.append(["Message"])  # เพิ่มหัวตาราง
    workbook.save(EXCEL_FILE)

@app.route('/')
def home():
    return '''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Save to Excel</title>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100vh;
                    background-color: #f4f4f9;
                }
                .container {
                    max-width: 400px;
                    text-align: center;
                    background: #fff;
                    padding: 20px 30px;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                    border-radius: 10px;
                    width: 90%; /* ทำให้เหมาะกับหน้าจอมือถือ */
                }
                h1 {
                    color: #333;
                    margin-bottom: 20px;
                    font-size: 1.5rem; /* ปรับขนาดสำหรับมือถือ */
                }
                form {
                    display: flex;
                    flex-direction: column;
                    gap: 15px;
                }
                label {
                    font-size: 1rem;
                    color: #555;
                }
                input[type="text"] {
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    font-size: 1rem;
                    width: 100%;
                    box-sizing: border-box;
                    outline: none;
                    transition: border-color 0.3s ease;
                }
                input[type="text"]:focus {
                    border-color: #007BFF;
                }
                button {
                    background-color: #007BFF;
                    color: #fff;
                    padding: 10px;
                    font-size: 1rem;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    transition: background-color 0.3s ease;
                }
                button:hover {
                    background-color: #0056b3;
                }
                .link {
                    margin-top: 20px;
                }
                a {
                    color: #007BFF;
                    text-decoration: none;
                    font-size: 0.9rem;
                }
                a:hover {
                    text-decoration: underline;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>มีอะไรจะบอกก่อนปีใหม่ไหม</h1>
                <form action="/save_to_excel" method="POST">
                    <label for="userInput">ใส่ข้อความ:</label>
                    <input type="text" id="userInput" name="userInput" placeholder="กรุณากรอกข้อความ" required>
                    <button type="submit">บันทึก</button>
                </form>
                <div class="link">
                    <a href="/download">ดาวน์โหลดไฟล์ Excel</a>
                </div>
            </div>
        </body>
        </html>
    '''

@app.route('/save_to_excel', methods=['POST'])
def save_to_excel():
    user_input = request.form['userInput']

    # บันทึกข้อความลง Excel
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([user_input])
    workbook.save(EXCEL_FILE)

    return redirect('/')

@app.route('/download')
def download_file():
    return send_file(EXCEL_FILE, as_attachment=True)

if __name__ == '__main__':
    # รันเซิร์ฟเวอร์ให้เข้าถึงได้ในเครือข่าย
    app.run(host='0.0.0.0', port=5000, debug=True)
